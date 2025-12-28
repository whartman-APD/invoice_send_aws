import logging
import math
import os
import time
import json
from datetime import datetime, timezone
import requests
import pandas
from dateutil.relativedelta import relativedelta
import io
import ast
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
import apd_quickbooksonline as quickbooks_online
import apd_msgraph_v2 as msgraph
import apd_clickup as clickup
import boto3
import apd_common
from dataclasses import dataclass

# Configuration (loaded from environment variables)
UPLOAD_TO_SHAREPOINT = os.environ.get("UPLOAD_TO_SHAREPOINT", "false").lower() == "true"
CREATE_INVOICE = os.environ.get("CREATE_INVOICE", "false").lower() == "true"
UPDATE_CLICKUP = os.environ.get("UPDATE_CLICKUP", "false").lower() == "true"
LOWER_CLIENT_ID = int(os.environ.get("LOWER_CLIENT_ID", "10000"))
UPPER_CLIENT_ID = int(os.environ.get("UPPER_CLIENT_ID", "20030"))
NET_30_DAYS_CLIENTS = os.environ.get("NET_30_DAYS_CLIENTS", "").split(",")

def get_billing_reference_date() -> datetime:
    """Get billing reference date from environment variable or default to first day of current month."""
    reference_date_str = os.environ.get("BILLING_REFERENCE_DATE", "")
    if reference_date_str:
        # Parse format: YYYY-MM-DD
        return datetime.strptime(reference_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    else:
        # Default to first day of current month
        now = datetime.now(timezone.utc)
        return datetime(now.year, now.month, 1, tzinfo=timezone.utc)


# Constants
SHAREPOINT_SITE_NAME = "APDClientFiles"
DOCUMENT_LIBRARY_NAME = "Documents"
APD_CLIENT_ID = "10000"
SUB_FOLDER_NAME = "Minutes"
BASE_PATH = f"10000 - Automata Practice Development/{SUB_FOLDER_NAME}"

@dataclass
class BillingPeriodConfig:
    """Configuration for billing periods."""
    reference_date: datetime
    
    @property
    def current_period_start(self) -> datetime:
        return self.reference_date
    
    @property
    def current_period_end(self) -> datetime:
        return (self.reference_date + relativedelta(months=1, days=-1)).replace(
            hour=23, minute=59, second=59
        )
    
    @property
    def prior_period_start(self) -> datetime:
        return self.reference_date - relativedelta(months=1)
    
    @property
    def prior_period_end(self) -> datetime:
        return (self.reference_date - relativedelta(days=1)).replace(
            hour=23, minute=59, second=59
        )
    
    @property
    def sharepoint_file_date(self) -> str:
        """Format: YYYY-M (e.g., '2025-9')"""
        return self.prior_period_start.strftime("%Y-%-m" if os.name != "nt" else "%Y-%#m")

# Initialize billing configuration
BILLING_CONFIG = BillingPeriodConfig(
    reference_date=get_billing_reference_date()
)

def process_all_clients():
    # Configure logging for Lambda/local execution
    logging.basicConfig(
        level=logging.INFO,
        format='%(levelname)s: %(message)s',
        force=True  # Force reconfiguration even if already configured
    )
    logging.info("Starting task minutes to ClickUp and QBO process...")
    
    aws_region = os.environ.get("AWS_REGION", "us-west-2")
    # Get the secrets from the vault
    try:
        aws_secretsmanager = boto3.client("secretsmanager", region_name=aws_region)
        quickbooks_online_vault = apd_common.get_secrets("QBO_SECRET_NAME", aws_secretsmanager)
        clickup_vault = apd_common.get_secrets("CLICKUP_SECRET_NAME", aws_secretsmanager)
        robocorp_vault = apd_common.get_secrets("ROBOCORP_API_SECRET_NAME", aws_secretsmanager)
        msgraph_vault = apd_common.get_secrets("MSGRAPH_SECRET_NAME", aws_secretsmanager)
        msgraph_instance = msgraph.MsGraph(
                tenant=msgraph_vault["tenant_id"],
                client_id=msgraph_vault["client_id"],
                client_secret=msgraph_vault["client_secret_value"],
                hostname=msgraph_vault["hostname"]
            )
        aws_dynamodb = boto3.resource('dynamodb', region_name=aws_region)
        client_orgs_table = apd_common.get_dynamodb_table("DYNAMODB_TABLE_ROBOCORP_CLIENTS", aws_dynamodb)
    except Exception as e:
        logging.error(f"Failed to initialize instances: {e}")
        return False

    unattended_data = get_unattended_data_from_sharepoint(msgraph_instance)
    
    for item in client_orgs_table.scan()['Items']:
        client_number = item['client_number']
        organization_id = item['organization_id']
        workspace_id = item['workspace_id']
        robocorp_control_room_api_key = robocorp_vault[client_number]
        
        logging.info(f"Processing client number: {client_number}")
        
        if not (LOWER_CLIENT_ID <= int(client_number) < UPPER_CLIENT_ID):
            continue
        
        if client_number == "10000":
            logging.info("Skipping Automata client number")
            continue

        header = {
        "Content-Type": "application/json",
        "Authorization": f"RC-WSKEY {robocorp_control_room_api_key}"
        }
        

        total_runtime_prior_month_unattended, unattended_export_file_stream, organization_name = get_unattended_data_from_spreadsheet(unattended_data, client_number, organization_id)

        total_runtime_prior_month_assistant, assistant_export_file_stream, dataframe_prior_month_assistant = get_assistant_runs(
            BILLING_CONFIG.prior_period_end,
            BILLING_CONFIG.prior_period_start,
            workspace_id,
            header,
            organization_name
        )

        dataframe_prior_months_unattended = get_unattended_runs(
            workspace_id,
            header,  
        )

        total_runtime_prior_month = total_runtime_prior_month_assistant + total_runtime_prior_month_unattended
        logging.info(f"Total runtime for client {client_number} for prior month: {total_runtime_prior_month} minutes")
        
        _, monthly_rate, included_minutes, consumption_rate, day_to_bill, service_type, client_type, billing_cc = (
            send_data_to_clickup(clickup_vault, client_number, total_runtime_prior_month)
        )
        
        report_datastream = build_runtime_report(client_number, dataframe_prior_months_unattended, dataframe_prior_month_assistant, included_minutes, consumption_rate)
        
        invoice_json = generate_invoice(
            quickbooks_online_vault,
            client_number,
            monthly_rate,
            included_minutes,
            consumption_rate,
            total_runtime_prior_month,
            day_to_bill,
            service_type,
            client_type,
            billing_cc,
        )

        if report_datastream and invoice_json:
            attach_detail_runtime_to_invoice(quickbooks_online_vault, invoice_json, report_datastream)

        send_files_to_sharepoint(
            msgraph_instance,
            client_number,
            assistant_export_file_stream,
            unattended_export_file_stream,
            report_datastream,
        )
        
        print("=====================================")
    return True

def send_files_to_sharepoint(msgraph_instance: msgraph.MsGraph, client_number: str, assistant_export_file_stream: str, unattended_export_file_stream: str, report_datastream: str):
    # Re-Authenticate and get access token
    msgraph_instance.access_token = msgraph_instance.request_access_token()
    # Get the site ID and drive ID for the Sharepoint site
    _, drive_id = get_site_id_and_drive_id(msgraph_instance, SHAREPOINT_SITE_NAME, DOCUMENT_LIBRARY_NAME)
    
    if UPLOAD_TO_SHAREPOINT:
        # Upload the files to the subfolder
        attended_export_filename = client_number + "_assistant_processes_" + BILLING_CONFIG.sharepoint_file_date + ".xlsx"
        msgraph_instance.upload_file_to_sharepoint(
                drive_id,
                BASE_PATH,
                attended_export_filename,
                assistant_export_file_stream,
                )    
        # Upload unattended processes file
        unattended_export_filename = client_number + "_unattended_processes_" + BILLING_CONFIG.sharepoint_file_date + ".xlsx"
        msgraph_instance.upload_file_to_sharepoint(
                drive_id,
                BASE_PATH,
                unattended_export_filename,
                unattended_export_file_stream,
            )
    
        # Upload the files to the subfolder
        report_filename = client_number + "_runtime_report_" + BILLING_CONFIG.sharepoint_file_date + ".pdf"
        msgraph_instance.upload_file_to_sharepoint(
                drive_id,
                BASE_PATH,
                report_filename,
                report_datastream,
            )

def attach_detail_runtime_to_invoice(quickbooks_online_vault: dict[str, str], invoice_json: dict[str, str]|None, report_datastream: str):
    # Get the invoice ID from the response
    if invoice_json is None:
        print("No invoice JSON returned. Skipping attachment.")
        return
    invoice_id = invoice_json["Invoice"]["Id"]
    print(f"Invoice ID: {invoice_id}")

    quickbooks_online_instance = quickbooks_online.QuickBooksOnline(quickbooks_online_vault)
    quickbooks_online_instance.upload_attachment(report_datastream, "invoice.pdf", "Invoice", invoice_id, content_type="application/pdf")

    print(f"Attached report to invoice {invoice_id} in QuickBooks Online.")

def generate_invoice(quickbooks_online_vault: dict[str, str], client_number: str, monthly_rate: float, included_minutes: int, consumption_rate: float, total_runtime_prior_month: int, day_to_bill: str, service_type: str, client_type: str, billing_cc: str):
    # Get the day to bill from the custom field
    current_month_and_year = datetime.now().replace(day=int(day_to_bill))
    formatted_date = current_month_and_year.strftime("%Y-%m-%d")
    due_date = formatted_date
    if client_number in NET_30_DAYS_CLIENTS:
        due_date = (current_month_and_year + relativedelta(months=1)).strftime("%Y-%m-%d")

    formatted_date_long = current_month_and_year.strftime("%B %d, %Y")
    next_billing_date_long = (current_month_and_year + relativedelta(months=1) - relativedelta(days=1)).strftime("%B %d, %Y")
    prior_month_and_year = (datetime.now() - relativedelta(months=1)).strftime("%B %Y")
    overage_description = f"Runtime Overage for {prior_month_and_year}"

    description = ""
    if service_type == "Managed Service" and client_type == "Client":
        description = f"Managed Automation Services for the period from {formatted_date_long} to {next_billing_date_long}."
    elif service_type == "Managed Service" and client_type == "Client (Maintenance)":
        description = f"Managed Automation Maintenance for the period from {formatted_date_long} to {next_billing_date_long}."

    # Managed Services Item ID
    managed_services_item_id = "11"
    managed_services_item_name = "Managed Automation Services"
    # Runtime Overage Item ID
    overage_item_id = "1010000001"
    overage_item_name = "Runtime Overage Minutes"

    # Get the customer information from QuickBooks Online
    query_string = f"SELECT * FROM Customer WHERE FullyQualifiedName LIKE'{client_number}%'"
    quickbooks_online_instance = quickbooks_online.QuickBooksOnline(quickbooks_online_vault)
    response = quickbooks_online_instance.query_a_customer(query_string)

    # Calculate the overage minutes
    if total_runtime_prior_month > included_minutes:
        overage_minutes = total_runtime_prior_month - included_minutes
    else:
        overage_minutes = 0
    
    # Create Invoice JSON
    line_items = []
    line_item = {
        "DetailType": "SalesItemLineDetail",
        "Amount": monthly_rate,
        "Description": description,
        "SalesItemLineDetail": {
            "ItemRef": {
                "value": managed_services_item_id,
                "name": managed_services_item_name,
            }
        },
    }
    line_items.append(line_item)
    if overage_minutes > 0:
        line_item = {
            "Description": overage_description, 
            "DetailType": "SalesItemLineDetail",
            "Amount": float(overage_minutes * consumption_rate),
            "SalesItemLineDetail": {
                "ItemRef": {"value": overage_item_id, "name": overage_item_name},
                "UnitPrice": consumption_rate,
                "Qty": int(overage_minutes),
            },
        }
        line_items.append(line_item)
    invoice = {
        "AllowIPNPayment": True,
        "AllowOnlineCreditCardPayment": False,
        "AllowOnlineACHPayment": True,
        "TxnDate": formatted_date,
        "DueDate": due_date,
        "Line": line_items,
        "CustomerRef": {"value": response["QueryResponse"]["Customer"][0]["Id"]},
        "BillEmail": {"Address": response["QueryResponse"]["Customer"][0]["PrimaryEmailAddr"]["Address"]},
        "SalesTermRef": {"value": "1"}
    }

    if billing_cc:
        invoice["BillEmailCc"] = {"Address": billing_cc}

    # Create the invoice in QuickBooks Online
    if CREATE_INVOICE:
        invoice_json = quickbooks_online_instance.create_invoice(invoice)
    else:
        invoice_json = None

    return invoice_json

def send_data_to_clickup(clickup_vault: dict[str, str], client_number: str, total_runtime_prior_month: int):
    print(f"Total runtime for {client_number} for prior month: {total_runtime_prior_month} minutes")

    list_id = clickup_vault["CRM_Business_List"]
    special_custom_field_id = clickup_vault["CRM_Business_List_Ac_Num_Query"] # This is the custom field id for "Account #" so we can filter using query params

    query_parameters={
            "custom_fields": [{
                "field_id": special_custom_field_id,
                "operator": "IS NOT NULL"
            }]
        }
    query_parameters["custom_fields"] = json.dumps(query_parameters["custom_fields"])

    tasks_list = clickup.get_tasks(
        clickup_vault,
        list_id,
        query_parameters=query_parameters
        )

    found_organization = False
    organization_task_id = None 
    robocorp_prior_usage_column_id = None
    robocorp_lifetime_usage_column_id = None
    robocorp_lifetime_usage = 0
    monthly_rate = 0
    included_minutes = 0
    consumption_rate = 0.50
    day_to_bill = 0
    service_type = None
    client_type = None
    billing_cc = None

    for organization in tasks_list:
        for custom_field in organization["custom_fields"]:
            if custom_field["name"] == "Account #" and custom_field["value"] == client_number:
                print(f'clickup task ID: {organization["id"]}, Client ID: {client_number}')
                print(f'Custom Field ID: {custom_field["id"]}')
                print(f'Organization Name: {organization["name"]}')
                organization_task_id = organization["id"]
                found_organization = True
                break
        if found_organization:
            for custom_field in organization["custom_fields"]:
                match custom_field["name"]:
                    case "Robocorp Prior Month":
                        robocorp_prior_usage_column_id = custom_field["id"]
                    case "Robocorp Lifetime":
                        robocorp_lifetime_usage_column_id = custom_field["id"]
                        robocorp_lifetime_usage = int(custom_field.get("value", 0))
                    case "Rate":
                        monthly_rate = int(custom_field.get("value", 0))
                    case "Included Consumption":
                        included_minutes = int(custom_field.get("value", 0))
                    case "Consumption Rate":
                        consumption_rate = float(custom_field.get("value", 0))
                    case "Day to Bill":
                        day_to_bill = int(custom_field.get("value", 0))
                    case "Service Type":
                        service_type_index = custom_field.get("value", "")
                        service_type = custom_field.get("type_config", "").get("options", [])[service_type_index].get("name", "")
                    case "Type":
                        client_type_index = custom_field.get("value", "")
                        client_type = custom_field.get("type_config", "").get("options", [])[client_type_index].get("name", "")
                    case "Billing CC":
                        billing_cc = custom_field.get("value", "")
                    case _:
                        pass #NOSONAR

            break

    if not found_organization:
        print(f"Organization with client ID {client_number} not found in ClickUp")
        # Send Email to Wes
        # continue

    # add total to the lifetime usage
    robocorp_lifetime_usage += total_runtime_prior_month
    if UPDATE_CLICKUP:
        clickup.set_custom_field_value(clickup_vault, organization_task_id, robocorp_lifetime_usage_column_id, str(robocorp_lifetime_usage))
    print(f"Total lifetime usage for {client_number}: {robocorp_lifetime_usage} minutes")

    # set prior month usage
    if UPDATE_CLICKUP:
        clickup.set_custom_field_value(clickup_vault, organization_task_id, robocorp_prior_usage_column_id, str(total_runtime_prior_month))

    return organization_task_id, monthly_rate, included_minutes, consumption_rate, day_to_bill, service_type, client_type, billing_cc

def get_unattended_data_from_spreadsheet(unattended_data:pandas.DataFrame, client_number:str, organization_id:str) -> tuple[int, io.BytesIO, str]:  
    # Filter the data for the Organization ID
    unattended_data_for_organization:pandas.DataFrame = unattended_data[unattended_data["Organization ID"] == organization_id]
    total_runtime_prior_month_unattended = unattended_data_for_organization['Process total run minutes used'].sum()
    
    #Remove all columns except those needed for export
    unattended_data_for_organization = unattended_data_for_organization[['Organization ID', 'Organization name', 'Process name', 'Process ID', 'Process total run minutes used', 'Process On-demand run minutes used']]
    
    export_file_stream = io.BytesIO()
    if unattended_data_for_organization.empty:
        print(f"No unattended processes found for {client_number}")
        organization_name = ""
        # Create an empty Excel writer object
        with pandas.ExcelWriter(export_file_stream, engine="openpyxl") as writer:
            empty_df = pandas.DataFrame(columns=["Organization ID", "Organization name", "Process name", "Process ID", "Process total run minutes used", "Process On-demand run minutes used"])
            empty_df.to_excel(writer, index=False, sheet_name="Unattended Processes")
    else:
        organization_name = unattended_data_for_organization['Organization name'].iloc[0]
        # Create an Excel writer object
        with pandas.ExcelWriter(export_file_stream, engine="openpyxl") as writer:
            unattended_data_for_organization.to_excel(writer, index=False, sheet_name="Unattended Processes")

    return total_runtime_prior_month_unattended, export_file_stream, organization_name

def get_unattended_runs(workspace_id: str, header: dict[str, str]) -> pandas.DataFrame:
    print("Getting Unattended Runs")
    url = f"https://cloud.robocorp.com/api/v1/workspaces/{workspace_id}/process-runs"
    unattended_process_list = []
    query_params = {
        "limit": 500,
    }
    count = 1
    while url:
        logging.info(f"Fetching unattended process runs page {count}")   
        count += 1
        response = requests.get(url, headers=header, params=query_params)
        response_json = response.json()
        unattended_process_list.extend(response_json.get('data', []))
        url = response_json.get('next') if response_json.get('has_more') else None

    # Convert all_data list to a DataFrame
    dataframe_unattended_processes = pandas.DataFrame(unattended_process_list)
    if dataframe_unattended_processes.empty:
        print("No unattended processes found.")
        return pandas.DataFrame()

    # Convert the 'started_at' column to a datetime object
    dataframe_unattended_processes['started_at'] = pandas.to_datetime(dataframe_unattended_processes['started_at'])

    # Initialize the 'runtime' column to 0
    dataframe_unattended_processes['runtime'] = 0
    
    # Filter the DataFrame for rows in the prior two month
    dataframe_prior_months_unattended = dataframe_unattended_processes[
        (dataframe_unattended_processes['started_at'] >= BILLING_CONFIG.prior_period_start) &
        (dataframe_unattended_processes['started_at'] <= BILLING_CONFIG.current_period_end)
    ]
    
    # Get the runtime for each process run in the filtered DataFrame using the step duration
    count = 1
    for index, row in dataframe_prior_months_unattended.iterrows():
        logging.info(f"Processing unattended run {count} of {len(dataframe_prior_months_unattended)}")
        count += 1
        process_id = row['id']
        query_params = {
            "process_run_id": process_id
        }
        url = f"https://cloud.robocorp.com/api/v1/workspaces/{workspace_id}/step-runs"
        unattended_run_list = []
        while url:
            try:
                response = requests.get(url, headers=header, params=query_params)
            except requests.exceptions.RequestException as e:
                print(f"Trying again after 15 seconds due to error: {e}")
                time.sleep(15)
                response = requests.get(url, headers=header, params=query_params)

            response_json = response.json()
            unattended_run_list.extend(response_json.get('data', []))
            url = response_json.get('next') if response_json.get('has_more') else None

        # Each step rounds up to the nearest minute and gets added to the total for the process run
        rounded_minutes = 0
        for step in unattended_run_list:
            if step['duration'] is not None:
                rounded_minutes = math.ceil(step['duration'] / 60) + rounded_minutes
        dataframe_prior_months_unattended.at[index, 'runtime'] = rounded_minutes

    return dataframe_prior_months_unattended

def build_runtime_report(client_number: str, dataframe_prior_months_unattended: pandas.DataFrame, dataframe_prior_month_assistant: pandas.DataFrame, included_minutes: int, consumption_rate: float):
    
    # check if empty. If not empty, Remove Columns, Rename Columns and merge together. If not, then create empty dataframe with correct columns
    if not dataframe_prior_month_assistant.empty:
        df_assistant_trimmed = dataframe_prior_month_assistant[["Process Name", "started_at", "runtime"]].copy()
        df_assistant_trimmed.rename(columns={"Process Name": "Process", "started_at": "Date", "runtime": "Runtime"}, inplace=True)
    else:
        df_assistant_trimmed = pandas.DataFrame(columns=["Process", "Date", "Runtime"])
    if not dataframe_prior_months_unattended.empty:
        df_unattended_trimmed = dataframe_prior_months_unattended[["process", "started_at", "runtime"]].copy()
        df_unattended_trimmed.rename(columns={"process": "Process", "started_at": "Date", "runtime": "Runtime"}, inplace=True)
    else:
        df_unattended_trimmed = pandas.DataFrame(columns=["Process", "Date", "Runtime"])
    df_trimmed = pandas.concat([df_assistant_trimmed, df_unattended_trimmed], ignore_index=True)
    
    if df_trimmed.empty:
        print("No data to build report.")
        return None
    
    # Clean up Names and Convert to DateTime
    def extract_process_name(x: dict[str, str] | str) -> str|None:
        return x.get("name", "") if isinstance(x, dict) else None
    df_trimmed["Process"] = df_trimmed["Process"].apply(extract_process_name)
    
    df_trimmed["Date"] = pandas.to_datetime(df_trimmed["Date"], utc=True).dt.tz_localize(None).dt.date
    pivot_table_df = df_trimmed.copy()
    bar_chart_df = df_trimmed.copy()
    run_data_df = df_trimmed.copy()

    pivot_table_df["Runtime"] = pandas.to_numeric(pivot_table_df["Runtime"], errors="coerce")
    pivot_table_df["Month"] = pandas.to_datetime(pivot_table_df["Date"]).dt.to_period("M").astype(str)

    pivot_table = pivot_table_df.pivot_table(
        index="Process",
        columns="Month",
        values="Runtime",
        aggfunc="sum",
        fill_value=0
    ).reset_index()
    pivot_table.loc["Total"] = pivot_table.sum(numeric_only=True)
    pivot_table.loc["Total", "Process"] = "Total"
    # Get the total from the rightmost column (last month)
    rightmost_col = pivot_table.columns[-1]
    total_previous_month_runtime_data = pivot_table.loc[pivot_table["Process"] == "Total", rightmost_col]
    total_prior_month = total_previous_month_runtime_data.values[0] if not total_previous_month_runtime_data.empty else 0
    
    print(f"Total from the prior month ({rightmost_col}): {total_prior_month}")

    # Prepare Bar Chart Data for the last two months
    bar_chart_df["Month"] = pandas.to_datetime(bar_chart_df["Date"]).dt.to_period("M").astype(str)
    bar_chart_df["Day"] = pandas.to_datetime(bar_chart_df["Date"]).dt.day
    daily_summary = bar_chart_df.groupby(["Day", "Month"])["Runtime"].sum().unstack(fill_value=0).reset_index()

    # Create a new Excel file with sheets
    report_data_stream = io.BytesIO()
    with pandas.ExcelWriter(report_data_stream, engine="openpyxl") as writer:
        pivot_table.to_excel(writer, index=False, sheet_name="Usage Pivot")
        daily_summary.to_excel(writer, index=False, sheet_name="Two Month Run Compare")
        run_data_df.to_excel(writer, index=False, sheet_name="Run Data")
    
    if total_prior_month > 0:
        add_overage_calculation_sheet(included_minutes, consumption_rate, total_prior_month, report_data_stream)
    report_data_stream.seek(0)
    
    final_data_stream = build_monthly_graph(daily_summary, report_data_stream)

    return final_data_stream

def add_overage_calculation_sheet(included_minutes: int, consumption_rate: float, total_prior_month: int, report_data_stream: io.BytesIO):
    report_data_stream.seek(0)
    wb = load_workbook(report_data_stream)
    wb.create_sheet("Overage Calculation")
    ws = wb["Overage Calculation"]
    last_row = 1
    ws.cell(row=last_row, column=1, value="Prior Month Total Runtime")
    ws.cell(row=last_row, column=2, value=total_prior_month)
    ws.cell(row=last_row + 1, column=1, value="Included Minutes")
    ws.cell(row=last_row + 1, column=2, value=included_minutes)
    ws.cell(row=last_row + 2, column=1, value="Overage Minutes")
    overage_minutes = total_prior_month - included_minutes if int(total_prior_month) > int(included_minutes) else 0
    ws.cell(row=last_row + 2, column=2, value=overage_minutes)
    ws.cell(row=last_row + 3, column=1, value="Consumption Rate")
    ws.cell(row=last_row + 3, column=2, value=consumption_rate)
    ws.cell(row=last_row + 4, column=1, value="Total Overage Cost")
    total_cost = overage_minutes * consumption_rate
    ws.cell(row=last_row + 4, column=2, value=total_cost)
    print(f"Prior Month Total Runtime: {total_prior_month}")
    print(f"Included Minutes: {included_minutes}")
    print(f"Overage Minutes: {overage_minutes}")
    print(f"Consumption Rate: {consumption_rate}")
    print(f"Total Overage Cost: {total_cost}")
    wb.save(report_data_stream)

def build_monthly_graph(daily_summary: pandas.DataFrame, report_data_stream: io.BytesIO):
    report_data_stream.seek(0)
    wb = load_workbook(report_data_stream)

    #Autosize columns in all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Excel column index (1-based)
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if cell_value:
                        max_length = max(max_length, len(cell_value))
                except:
                    pass
            adjusted_width = max_length + 2  # Add a little extra padding
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Bold the "Total" row in the "Usage Pivot" sheet
    ws_pivot = wb["Usage Pivot"]
    for cell in ws_pivot.iter_rows(min_row=ws_pivot.max_row, max_row=ws_pivot.max_row):
        for c in cell:
            c.font = Font(bold=True)
            c.border = Border(top=Side(style="thin"))

    # Add a Bar Chart
    ws_graph = wb["Two Month Run Compare"]
    chart = BarChart()
    chart.x_axis.title = "Day of Month"
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.delete = False
    chart.x_axis.tickLblSkip = 7
    chart.y_axis.title = "Runtime (minutes)"
    chart.y_axis.majorTickMark = "in"
    chart.y_axis.delete = False
    chart.legend.position = "b"
    
    # Adjust chart size to fit the data and legends
    chart.width = 20  # Increase chart width
    chart.height = 10  # Increase chart height
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.05,
            y=0.05,
            h=0.8,
            w=0.95,
            xMode='edge',
            yMode='edge',
            hMode='edge',
            wMode='edge',
        )
    )

    # Set axis scaling
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = None  # Auto-adjust max based on data

    # Add data to the chart
    data = Reference(ws_graph, min_col=2, min_row=1, max_col=1 + len(daily_summary.columns) - 1, max_row=1 + len(daily_summary))
    cats = Reference(ws_graph, min_col=1, min_row=2, max_row=1 + len(daily_summary))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_graph.add_chart(chart)
    
    # Save the workbook to a stream
    final_stream = io.BytesIO()
    wb.save(final_stream)
    final_stream.seek(0)

    return final_stream

def get_assistant_runs(last_day_of_prior_month: str, first_day_of_prior_month: str, workspace_id: str, header: dict[str, str], organization_name: str) -> tuple[int, io.BytesIO, pandas.DataFrame]:
    print("Getting Assistant Runs")
    excel_stream = io.BytesIO()
    query_params = {
        "limit": 500,
    }
    url = f"https://cloud.robocorp.com/api/v1/workspaces/{workspace_id}/assistant-runs"
    assistant_run_list = []
    while url:
        response = requests.get(url, headers=header, params=query_params)
        response_json = response.json()
        assistant_run_list.extend(response_json.get('data', []))
        url = response_json.get('next') if response_json.get('has_more') else None

    # Convert all_data list to a DataFrame
    dataframe_assistant_runs = pandas.DataFrame(assistant_run_list)
    
    if dataframe_assistant_runs.empty:
        print("No Assistant runs found.")
        return 0, excel_stream, pandas.DataFrame()

    # Extract 'id' and 'name' from the 'assistant' column and create new columns
    def extract_process_info(x: dict[str, str] | str) -> tuple[str|None, str|None]:
        if isinstance(x, dict):
            return x.get("id", None), x.get("name", None)
        return None, None
    dataframe_assistant_runs['Process ID'], dataframe_assistant_runs['Process Name'] = zip(*dataframe_assistant_runs['assistant'].apply(extract_process_info))
    
    # Initialize columns
    dataframe_assistant_runs['Organization name'] = organization_name
    dataframe_assistant_runs['Organization ID'] = workspace_id
    dataframe_assistant_runs['started_at'] = pandas.to_datetime(dataframe_assistant_runs['started_at'])

    # Filter the DataFrame for rows in the prior month and with state "completed"
    dataframe_prior_month_assistant = dataframe_assistant_runs[
        (dataframe_assistant_runs['started_at'] >= first_day_of_prior_month) &
        (dataframe_assistant_runs['started_at'] <= last_day_of_prior_month)
    ].copy()

    if dataframe_prior_month_assistant.empty:
        print("No Assistant runs found for the prior month.")
        return 0, excel_stream, pandas.DataFrame()

    # Round the durations up to the nearest minute and assign it to the Process total run minutes used column
    def round_up_to_minute(x: int) -> int:
        return math.ceil(x / 60)
    rounded_minutes = dataframe_prior_month_assistant['duration'].apply(round_up_to_minute)
    dataframe_prior_month_assistant['Process total run minutes used'] = rounded_minutes
    dataframe_prior_month_assistant['runtime'] = rounded_minutes

    # Sum the Process total run minutes used for these filtered rows
    total_runtime_prior_month_assistant = dataframe_prior_month_assistant['Process total run minutes used'].sum()

    # Convert timezone-aware datetime columns to naive datetime
    for col in dataframe_prior_month_assistant.columns:
        if pandas.api.types.is_datetime64_any_dtype(dataframe_prior_month_assistant[col]):
            series = dataframe_prior_month_assistant[col]
            if hasattr(series.dt, 'tz') and series.dt.tz is not None:  # type: ignore
                dataframe_prior_month_assistant[col] = series.dt.tz_localize(None)  # type: ignore


    with pandas.ExcelWriter(excel_stream, engine='xlsxwriter') as writer:
        dataframe_prior_month_assistant.to_excel(writer, index=False, sheet_name='Assistant Runs')

    excel_stream.seek(0)    
    return total_runtime_prior_month_assistant, excel_stream, dataframe_prior_month_assistant

def get_site_id_and_drive_id(msgraph_instance: msgraph.MsGraph, site_name: str, document_library_name: str):
    #Sharepoint navigation
    response = msgraph_instance.get_sharepoint_site(site_name)
    site_id = response.json().get("id")
    response = msgraph_instance.get_sharepoint_drives(site_id)
    drive_id = msgraph_instance.get_drive_id_by_name(response.json(), document_library_name)
    return site_id, drive_id

def get_unattended_data_from_sharepoint(msgraph_instance:msgraph.MsGraph) -> pandas.DataFrame:
    unattended_spreadsheet = f'account-usage-a4db96d0-2dbb-481e-b35a-4629ff252457-{BILLING_CONFIG.sharepoint_file_date}.csv'
    
    # Get the site ID and drive ID for the Sharepoint site
    _, drive_id = get_site_id_and_drive_id(msgraph_instance, SHAREPOINT_SITE_NAME, DOCUMENT_LIBRARY_NAME)
    
    folder_list_json = msgraph_instance.get_folders_in_drive(drive_id)
    _, folder_id = msgraph_instance.get_item_name_starts_with(folder_list_json, APD_CLIENT_ID) # Find the client ID folder
    sub_folder_list_json = msgraph_instance.get_items_in_folder(drive_id, folder_id)
    _, sub_folder_id = msgraph_instance.get_item_name_starts_with(sub_folder_list_json, SUB_FOLDER_NAME) # Find the Minutes folder
    sub_items_list_json = msgraph_instance.get_items_in_folder(drive_id, sub_folder_id)
    _, sub_folder_id = msgraph_instance.get_item_name_starts_with(sub_items_list_json, "CSV Data") # Find the Minutes folder
    sub_items_list_json = msgraph_instance.get_items_in_folder(drive_id, sub_folder_id)
    
    unattended_file = next((item for item in sub_items_list_json if item["name"] == unattended_spreadsheet), None)
    
    if unattended_file:
        download_url = unattended_file["@microsoft.graph.downloadUrl"]
        response = requests.get(download_url)
        file_content = io.StringIO(response.content.decode('utf-8'))
        unattended_data = pandas.read_csv(file_content)
    else:
        print(f"File {unattended_spreadsheet} not found in SharePoint.")
        raise Exception(f"File {unattended_spreadsheet} not found in SharePoint.")
    
    return unattended_data

if __name__ == "__main__":
    process_all_clients()